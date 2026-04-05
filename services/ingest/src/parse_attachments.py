import io
import logging
from typing import Optional

logger = logging.getLogger(__name__)


async def parse_attachments(files: list[dict]) -> list[str]:
    """Parse Excel and PDF attachments, return text content from each."""
    results = []

    for f in files:
        filename = f.get("filename", "")
        content = f.get("content", b"")
        content_type = f.get("content_type", "")

        try:
            if filename.lower().endswith((".xlsx", ".xls")):
                text = parse_excel(content, filename)
                if text:
                    results.append(f"[Attachment: {filename}]\n{text}")

            elif filename.lower().endswith(".pdf") or "pdf" in content_type:
                text = parse_pdf(content, filename)
                if text:
                    results.append(f"[Attachment: {filename}]\n{text}")

            else:
                logger.info(f"Skipping unsupported attachment: {filename} ({content_type})")

        except Exception as e:
            logger.error(f"Failed to parse attachment {filename}: {e}")

    return results


def parse_excel(content: bytes, filename: str) -> Optional[str]:
    """Parse Excel file to text representation."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    lines = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            if any(v.strip() for v in vals):
                lines.append(" | ".join(vals))

    return "\n".join(lines) if lines else None


def parse_pdf(content: bytes, filename: str) -> Optional[str]:
    """Parse PDF file to text."""
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    lines = []

    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and text.strip():
            lines.append(f"--- Page {i + 1} ---")
            lines.append(text.strip())

    return "\n".join(lines) if lines else None
